#!/usr/bin/env python
# coding: utf-8

# # Download packages

# In[1]:


#Packages needed for the automation process
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import requests
import bs4
from bs4 import BeautifulSoup
import os, random, sys, time
from urllib.parse import urlparse
from selenium import webdriver
import pandas as pd
from openpyxl import load_workbook, Workbook
import urllib
import numpy as np


# # Set up browser and log in

# In[3]:


#setting chromebrowser after downloading a compatible version of the same and setting a path to that .exe file
browser = webdriver.Chrome("C:/Users/Admin/Automation/chromedriver.exe")


# In[18]:


#Opening automation chrome browser window
browser.get("https://www.linkedin.com/uas/login")


# In[19]:


#create a .txt file with username/emailID in the first line and password in the second
file = open("C:/Users/Admin/Automation/config.txt")
line = file.readlines()
username = line[0]
password = line[1]


# In[20]:


#pasting the login credentials and submitting the info to log into Linkedin
elementID = browser.find_element(by= By.ID, value="username")
elementID.send_keys(username)
elementID = browser.find_element(by = By.ID, value="password")
elementID.send_keys(password)
elementID.submit()


# # Create DataFrame and add profiles links

# In[22]:


#Creating a dataframe to store the info
df = pd.DataFrame(columns=["Company","Address","Telephone","Name","Job Title","Bio","Profile_Link"])
err_profile = [] 


# In[4]:


#put all the links together and press enter to let y record it.
profiles = input(str("Please enter profiles to be extracted: "))
profiles = profiles.split(" ")
profiles


# In[ ]:


total_errs=0
for full_link in profiles:
    total_errs = total_errs+1
    biodata = ""
    if full_link in list(df["Profile_Link"]):
        print("Profile ",total_errs," already exists.")
    else:
        browser.get(full_link)
        SCROLL_PAUSE_TIME = 4
        #wait to load page
        time.sleep(SCROLL_PAUSE_TIME)
        #Get scroll height
        last_height = browser.execute_script("return document.body.scrollHeight")

        for i in range(3):
            #scroll down to bottom
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            #Calculate new scroll height and compare with last scroll height
            new_height = browser.execute_script("return document.body.scrollHeight")
            last_height = new_height
        
        soup = BeautifulSoup(browser.page_source,"html.parser")
        name_div= soup.find("div",{"class":"ph5 pb5"})
        try:
            #Other info
            profile_name = name_div.find_all("h1")[0].get_text()
            profile_loc = name_div.find_all("span", class_ = "text-body-small inline t-black--light break-words")[0].get_text().strip()
            name_div2= soup.find_all("div",{"class":"pvs-list__outer-container"})[1]
            profile_jobtitle = name_div2.find_all("span")[1].get_text().strip()
            profile_company = name_div2.find_all("span")[5].get_text().strip()

            #Biodata
            
            profile_experience = full_link + "/details/experience/"
            browser.get(profile_experience)
            SCROLL_PAUSE_TIME = 4
            #wait to load page
            time.sleep(SCROLL_PAUSE_TIME)
            soup = BeautifulSoup(browser.page_source,"html.parser")
            
            
            for k in range(0,10):
                for j in range(0,7):
                    try:
                        name_div3= soup.find_all("div",{"class":"pvs-entity pvs-entity--padded pvs-list__item--no-padding-when-nested"})[k]
                        biodata = biodata + "\n"+(name_div3.find_all("span",{"class":"visually-hidden"}))[j].get_text().strip()[0:int(np.where((name_div3.find_all("span",{"class":"visually-hidden"})[j].get_text().strip().find("·")-1)<0,50,(name_div3.find_all("span",{"class":"visually-hidden"})[j].get_text().strip().find("·")-1)))]
                    except Exception:
                        continue

            df = df.append({"Bio":biodata,"Address":profile_loc,"Company":profile_company,"Name":profile_name,"Job Title":profile_jobtitle,"Profile_Link":full_link}, ignore_index = True)
        except Exception:
            err_profile.append(str(full_link))
            print("ERROR! Failed to extract profile number ", total_errs,":",full_link)
        SCROLL_PAUSE_TIME = 2
        time.sleep(SCROLL_PAUSE_TIME)
print(err_profile)


# # Saving the data

# In[ ]:


#Save the dataframe in an excel file
df.to_excel("C:/Users/Admin/Automation/profiles.xlsx")

#Copy the data from one excel to another one that already has the desired format
wb = load_workbook("C:/Users/Admin/Automation/profiles.xlsx")
wb2 = load_workbook("C:/Users/Admin/Automation/profiles_final.xlsx")
sheet1 = wb.get_sheet_by_name("Sheet1")
sheet2 = wb2.get_sheet_by_name("Sheet1")
sheet2.delete_cols(idx=1, amount = 10)

wb2.save("C:/Users/Admin/Automation/profiles_final.xlsx")

for rownumber in range(1,sheet1.max_row+1):
    for columnnumber in range(1,sheet1.max_column+1):
        sheet2.cell(row=rownumber,column=columnnumber).value = sheet1.cell(row=rownumber,column=columnnumber).value

wb.save("C:/Users/Admin/Automation/profiles.xlsx")
wb2.save("C:/Users/Admin/Automation/profiles_final.xlsx")


# # <center> ---------------------THE END--------------------- </center>
